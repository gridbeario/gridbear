"""Chat history REST API for user portal (PostgreSQL).

Provides conversation management and message persistence for WebChat.
Storage: PostgreSQL chat.webchat_conversations / chat.webchat_messages.
"""

import json
import os
import re
import time
import uuid
from pathlib import Path

from fastapi import APIRouter, Depends, Request, UploadFile
from fastapi.responses import FileResponse
from starlette.background import BackgroundTask

from config.logging_config import logger
from core.api_schemas import ApiResponse, api_error, api_ok
from core.encryption import decrypt, encrypt, is_encrypted
from ui.routes.auth import require_user

router = APIRouter(prefix="/me/chat/api", tags=["chat-api"])

BASE_DIR = Path(__file__).resolve().parent.parent.parent
ATTACHMENTS_DIR = BASE_DIR / "data" / "attachments"

# Max upload: 20MB per file, images + common doc types
MAX_UPLOAD_SIZE = 20 * 1024 * 1024
ALLOWED_EXTENSIONS = {
    ".jpg",
    ".jpeg",
    ".png",
    ".gif",
    ".webp",
    ".bmp",
    ".svg",
    ".pdf",
    ".txt",
    ".csv",
    ".json",
    ".xml",
    ".md",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".mp3",
    ".wav",
    ".m4a",
    ".ogg",
    ".flac",
    ".webm",
    ".mp4",
}

MIGRATION_NAME = "004_apps_webchat"
MIGRATION_SQL = BASE_DIR / "scripts" / "migrations" / f"{MIGRATION_NAME}.sql"

_db = None
_initialized = False


def _ensure_db():
    """Initialize PG database reference and apply migration if needed."""
    global _db, _initialized
    if _initialized:
        return

    from core.registry import get_database

    _db = get_database()
    if _db is None:
        raise RuntimeError("PostgreSQL database not available (chat API requires it)")

    with _db.acquire_sync() as conn:
        row = conn.execute(
            "SELECT 1 FROM public._migrations WHERE name = %s",
            (MIGRATION_NAME,),
        ).fetchone()
        if not row:
            sql = MIGRATION_SQL.read_text()
            conn.execute(sql)
            conn.execute(
                "INSERT INTO public._migrations (name) VALUES (%s)",
                (MIGRATION_NAME,),
            )
            conn.commit()
            logger.info(f"Applied {MIGRATION_NAME} migration (chat_api)")
        else:
            conn.rollback()

    # Ensure context_prompt column exists (idempotent)
    with _db.acquire_sync() as conn:
        conn.execute(
            "ALTER TABLE chat.webchat_conversations "
            "ADD COLUMN IF NOT EXISTS context_prompt TEXT"
        )
        # Shared conversations schema
        conn.execute(
            "ALTER TABLE chat.webchat_conversations "
            "ADD COLUMN IF NOT EXISTS type TEXT DEFAULT 'private'"
        )
        conn.execute(
            "ALTER TABLE chat.webchat_messages ADD COLUMN IF NOT EXISTS sender_id TEXT"
        )
        conn.execute("""
            CREATE TABLE IF NOT EXISTS chat.webchat_participants (
                conversation_id TEXT NOT NULL
                    REFERENCES chat.webchat_conversations(id) ON DELETE CASCADE,
                unified_id TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'member',
                joined_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (conversation_id, unified_id)
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_webchat_participants_uid
            ON chat.webchat_participants(unified_id)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS chat.webchat_invites (
                token TEXT PRIMARY KEY,
                conversation_id TEXT NOT NULL
                    REFERENCES chat.webchat_conversations(id) ON DELETE CASCADE,
                created_by TEXT NOT NULL,
                created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMPTZ NOT NULL,
                max_uses INTEGER DEFAULT 0,
                use_count INTEGER DEFAULT 0
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_webchat_invites_conv
            ON chat.webchat_invites(conversation_id)
        """)
        # Backfill: create owner participant for existing conversations
        conn.execute("""
            INSERT INTO chat.webchat_participants (conversation_id, unified_id, role)
            SELECT id, unified_id, 'owner'
            FROM chat.webchat_conversations
            WHERE id NOT IN (
                SELECT conversation_id FROM chat.webchat_participants
            )
            ON CONFLICT DO NOTHING
        """)
        # Conversation documents (per-conversation knowledge base)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS chat.webchat_documents (
                id TEXT PRIMARY KEY,
                conversation_id TEXT NOT NULL
                    REFERENCES chat.webchat_conversations(id) ON DELETE CASCADE,
                filename TEXT NOT NULL,
                original_filename TEXT NOT NULL,
                file_path TEXT NOT NULL,
                file_size INTEGER NOT NULL,
                mime_type TEXT,
                content_text TEXT,
                uploaded_by TEXT NOT NULL,
                uploaded_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_webchat_documents_conv
            ON chat.webchat_documents(conversation_id)
        """)
        conn.commit()

    # Plan tables migration
    with _db.acquire_sync() as conn:
        row = conn.execute(
            "SELECT 1 FROM public._migrations WHERE name = %s",
            ("010_webchat_plans",),
        ).fetchone()
        if not row:
            sql = (
                BASE_DIR / "scripts" / "migrations" / "010_webchat_plans.sql"
            ).read_text()
            conn.execute(sql)
            conn.execute(
                "INSERT INTO public._migrations (name) VALUES (%s)",
                ("010_webchat_plans",),
            )
            conn.commit()
            logger.info("Applied 010_webchat_plans migration")
        else:
            conn.rollback()

    _initialized = True


def _extract_text(file_path: str, mime_type: str | None) -> str:
    """Extract text content from a document for RAG indexing."""
    path = Path(file_path)
    ext = path.suffix.lower()
    try:
        if ext in (".txt", ".md", ".csv", ".json", ".xml"):
            return path.read_text(errors="replace")[:200_000]

        if ext == ".pdf":
            from pypdf import PdfReader

            reader = PdfReader(path)
            pages = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n\n".join(pages)[:200_000]

        if ext == ".docx":
            from docx import Document as DocxDocument

            doc = DocxDocument(path)
            return "\n\n".join(p.text for p in doc.paragraphs if p.text)[:200_000]

        if ext in (".xlsx", ".xls"):
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"[Sheet: {sheet}]")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    lines.append("\t".join(cells))
            wb.close()
            return "\n".join(lines)[:200_000]

    except Exception as e:
        logger.warning(f"Text extraction failed for {path.name}: {e}")
    return ""


def _uid(user: dict) -> str:
    return user["username"]


def _serialize_row(row: dict) -> dict:
    """Convert datetime values to ISO strings for JSON serialization."""
    return {k: v.isoformat() if hasattr(v, "isoformat") else v for k, v in row.items()}


# --- Public helpers (used by ws_chat) ---


def save_message(
    conversation_id: str,
    role: str,
    content: str,
    metadata: dict | None = None,
    sender_id: str | None = None,
) -> int | None:
    """Save a message and update conversation timestamp. Auto-title if empty.

    Returns the new ``chat.webchat_messages.id`` so callers can echo it
    back to the client over WebSocket — the frontend uses it to backfill
    ``msg.dbId`` on locally-pushed messages, which is the prerequisite
    for the per-message DELETE endpoint to address the right row.
    """
    _ensure_db()
    encrypted_content = encrypt(content)
    new_id: int | None = None
    with _db.acquire_sync() as conn:
        cur = conn.execute(
            """INSERT INTO chat.webchat_messages
               (conversation_id, role, content, metadata_json, sender_id)
               VALUES (%s, %s, %s, %s, %s)
               RETURNING id""",
            (
                conversation_id,
                role,
                encrypted_content,
                json.dumps(metadata) if metadata else None,
                sender_id,
            ),
        )
        row = cur.fetchone()
        if row:
            new_id = row["id"]
        conn.execute(
            "UPDATE chat.webchat_conversations SET updated_at = CURRENT_TIMESTAMP WHERE id = %s",
            (conversation_id,),
        )
        # Auto-title: set title from first user message
        if role == "user":
            row = conn.execute(
                "SELECT title FROM chat.webchat_conversations WHERE id = %s",
                (conversation_id,),
            ).fetchone()
            if row and not row["title"]:
                title = content[:80].strip()
                if len(content) > 80:
                    title += "..."
                conn.execute(
                    "UPDATE chat.webchat_conversations SET title = %s WHERE id = %s",
                    (title, conversation_id),
                )
        conn.commit()
    return new_id


def get_conversation_title(conversation_id: str) -> str:
    """Get current title of a conversation."""
    _ensure_db()
    with _db.acquire_sync() as conn:
        row = conn.execute(
            "SELECT title FROM chat.webchat_conversations WHERE id = %s",
            (conversation_id,),
        ).fetchone()
    return row["title"] if row else ""


def validate_conversation_ownership(conversation_id: str, unified_id: str) -> bool:
    """Check that a conversation belongs to the given user (legacy — private only)."""
    _ensure_db()
    with _db.acquire_sync() as conn:
        row = conn.execute(
            "SELECT 1 FROM chat.webchat_conversations WHERE id = %s AND unified_id = %s",
            (conversation_id, unified_id),
        ).fetchone()
    return row is not None


def validate_conversation_access(conversation_id: str, uid: str) -> str | None:
    """Check if user is a participant (owner or member). Returns role or None."""
    _ensure_db()
    with _db.acquire_sync() as conn:
        row = conn.execute(
            "SELECT role FROM chat.webchat_participants "
            "WHERE conversation_id = %s AND unified_id = %s",
            (conversation_id, uid),
        ).fetchone()
        return row["role"] if row else None


def is_conversation_owner(conversation_id: str, uid: str) -> bool:
    """Check if user is the owner of a conversation."""
    return validate_conversation_access(conversation_id, uid) == "owner"


def list_conversation_participants(conversation_id: str) -> list[str]:
    """Return list of participant unified_ids for a conversation."""
    _ensure_db()
    with _db.acquire_sync() as conn:
        rows = conn.execute(
            "SELECT unified_id FROM chat.webchat_participants "
            "WHERE conversation_id = %s",
            (conversation_id,),
        ).fetchall()
    return [r["unified_id"] for r in rows]


# --- REST endpoints ---


@router.get(
    "/conversations",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def list_conversations(request: Request, user: dict = Depends(require_user)):
    _ensure_db()
    agent = request.query_params.get("agent", "")
    uid = _uid(user)
    with _db.acquire_sync() as conn:
        if agent:
            cur = conn.execute(
                """SELECT c.id, c.agent_name, c.title, c.created_at,
                          c.updated_at, c.type, c.context_prompt,
                          p.pinned_at,
                          EXISTS(
                              SELECT 1 FROM chat.webchat_plans pl
                              WHERE pl.conversation_id = c.id
                              AND pl.status NOT IN ('completed', 'cancelled')
                          ) as has_plan
                   FROM chat.webchat_conversations c
                   JOIN chat.webchat_participants p
                        ON c.id = p.conversation_id
                   WHERE p.unified_id = %s AND c.agent_name = %s
                   ORDER BY p.pinned_at IS NULL, p.pinned_at ASC,
                            c.updated_at DESC""",
                (uid, agent),
            )
        else:
            cur = conn.execute(
                """SELECT c.id, c.agent_name, c.title, c.created_at,
                          c.updated_at, c.type, c.context_prompt,
                          p.pinned_at,
                          EXISTS(
                              SELECT 1 FROM chat.webchat_plans pl
                              WHERE pl.conversation_id = c.id
                              AND pl.status NOT IN ('completed', 'cancelled')
                          ) as has_plan
                   FROM chat.webchat_conversations c
                   JOIN chat.webchat_participants p
                        ON c.id = p.conversation_id
                   WHERE p.unified_id = %s
                   ORDER BY p.pinned_at IS NULL, p.pinned_at ASC,
                            c.updated_at DESC""",
                (uid,),
            )
        rows = []
        for r in cur.fetchall():
            row = _serialize_row(dict(r))
            row["pinned"] = row.get("pinned_at") is not None
            rows.append(row)
    return api_ok(data={"conversations": rows})


@router.post(
    "/conversations",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def create_conversation(request: Request, user: dict = Depends(require_user)):
    try:
        body = await request.json()
    except Exception:
        return api_error(400, "Invalid JSON", "validation_error")

    agent_name = body.get("agent_name", "")
    uid = _uid(user)

    # Validate agent access
    if agent_name and not user.get("is_superadmin"):
        from ui.routes.me import _get_allowed_users
        from ui.routes.ws_chat import _load_agent_yaml

        agent_cfg = _load_agent_yaml(agent_name)
        if agent_cfg:
            allowed = _get_allowed_users(agent_cfg)
            if not allowed or uid.lower() not in allowed:
                return api_error(403, "Access denied to this agent", "forbidden")

    conv_id = str(uuid.uuid4())

    _ensure_db()
    with _db.acquire_sync() as conn:
        conn.execute(
            """INSERT INTO chat.webchat_conversations (id, unified_id, agent_name)
               VALUES (%s, %s, %s)""",
            (conv_id, uid, agent_name),
        )
        # Create owner participant
        conn.execute(
            """INSERT INTO chat.webchat_participants
               (conversation_id, unified_id, role)
               VALUES (%s, %s, 'owner')""",
            (conv_id, uid),
        )
        conn.commit()
        row = conn.execute(
            """SELECT id, agent_name, title, created_at, updated_at, type
               FROM chat.webchat_conversations WHERE id = %s""",
            (conv_id,),
        ).fetchone()
    return api_ok(data={"conversation": _serialize_row(dict(row))})


@router.get(
    "/conversations/{conv_id}/messages",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def get_messages(
    request: Request, conv_id: str, user: dict = Depends(require_user)
):
    _ensure_db()
    uid = _uid(user)
    if not validate_conversation_access(conv_id, uid):
        return api_error(403, "Access denied", "forbidden")

    limit = int(request.query_params.get("limit", "200"))
    offset = int(request.query_params.get("offset", "0"))

    with _db.acquire_sync() as conn:
        cur = conn.execute(
            """SELECT * FROM (
                   SELECT m.id, m.role, m.content, m.metadata_json, m.created_at,
                          m.sender_id, u.display_name as sender_display_name
                   FROM chat.webchat_messages m
                   LEFT JOIN app.users u ON u.username = m.sender_id
                   WHERE m.conversation_id = %s
                   ORDER BY m.created_at DESC
                   LIMIT %s OFFSET %s
               ) sub ORDER BY sub.created_at ASC""",
            (conv_id, limit, offset),
        )
        rows = []
        for r in cur.fetchall():
            msg = _serialize_row(dict(r))
            # Decrypt content (handles both encrypted and pre-migration plaintext)
            raw = msg.get("content", "")
            msg["content"] = decrypt(raw) if is_encrypted(raw) else raw
            if msg["metadata_json"]:
                msg["metadata"] = json.loads(msg["metadata_json"])
            del msg["metadata_json"]
            rows.append(msg)

    return api_ok(data={"messages": rows})


@router.delete(
    "/conversations/{conv_id}/messages/{msg_id}",
    response_model=ApiResponse,
    response_model_exclude_none=True,
)
async def delete_message(
    request: Request,
    conv_id: str,
    msg_id: int,
    user: dict = Depends(require_user),
):
    """Hard-delete a single webchat message. Superadmin only.

    Writes an entry to admin.audit_log with role + content length but NOT
    the content itself — the column is encrypted at rest and the typical
    reason for deletion is sensitive data, so leaking the plaintext into
    a separate log defeats the purpose. Use a backup if recovery is
    needed.
    """
    if not user.get("is_superadmin"):
        return api_error(403, "Superadmin only", "forbidden")

    _ensure_db()
    with _db.acquire_sync() as conn:
        row = conn.execute(
            "SELECT role, length(content) AS content_len "
            "FROM chat.webchat_messages "
            "WHERE id = %s AND conversation_id = %s",
            (msg_id, conv_id),
        ).fetchone()
        if not row:
            return api_error(404, "Not found", "not_found")
        conn.execute(
            "DELETE FROM chat.webchat_messages WHERE id = %s",
            (msg_id,),
        )
        conn.commit()

    try:
        from ui.auth.database import AuthDatabase

        AuthDatabase().log_event(
            event_type="webchat_message_delete",
            user_id=user.get("id"),
            username=user.get("username"),
            ip_address=request.client.host if request.client else None,
            success=True,
            details=json.dumps(
                {
                    "conv_id": conv_id,
                    "msg_id": msg_id,
                    "role": row["role"],
                    "content_length": row["content_len"],
                }
            ),
        )
    except Exception as exc:
        # Audit-log failure must not roll back the deletion (already
        # committed) but should be visible to operators.
        logger.warning("audit log for webchat_message_delete failed: %s", exc)

    return api_ok()


@router.post(
    "/conversations/{conv_id}/rename",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def rename_conversation(
    request: Request, conv_id: str, user: dict = Depends(require_user)
):
    try:
        body = await request.json()
    except Exception:
        return api_error(400, "Invalid JSON", "validation_error")

    title = (body.get("title") or "").strip()
    if not title:
        return api_error(400, "Title required", "validation_error")

    _ensure_db()
    uid = _uid(user)
    with _db.acquire_sync() as conn:
        row = conn.execute(
            "SELECT 1 FROM chat.webchat_conversations WHERE id = %s AND unified_id = %s",
            (conv_id, uid),
        ).fetchone()
        if not row:
            return api_error(404, "Not found", "not_found")

        conn.execute(
            "UPDATE chat.webchat_conversations SET title = %s WHERE id = %s",
            (title, conv_id),
        )
        conn.commit()
    return api_ok(data={"title": title})


@router.delete(
    "/conversations/{conv_id}",
    response_model=ApiResponse,
    response_model_exclude_none=True,
)
async def delete_conversation(
    request: Request, conv_id: str, user: dict = Depends(require_user)
):
    _ensure_db()
    uid = _uid(user)
    with _db.acquire_sync() as conn:
        row = conn.execute(
            "SELECT 1 FROM chat.webchat_conversations WHERE id = %s AND unified_id = %s",
            (conv_id, uid),
        ).fetchone()
        if not row:
            return api_error(404, "Not found", "not_found")

        conn.execute(
            "DELETE FROM chat.webchat_conversations WHERE id = %s",
            (conv_id,),
        )
        conn.commit()
    return api_ok()


# --- Conversation context ---


@router.get(
    "/conversations/{conv_id}/context",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def get_context(
    request: Request, conv_id: str, user: dict = Depends(require_user)
):
    """Get conversation context prompt."""
    _ensure_db()
    uid = _uid(user)
    with _db.acquire_sync() as conn:
        row = conn.execute(
            "SELECT context_prompt FROM chat.webchat_conversations "
            "WHERE id = %s AND unified_id = %s",
            (conv_id, uid),
        ).fetchone()
        if not row:
            return api_error(404, "Not found", "not_found")
    return api_ok(data={"context_prompt": row["context_prompt"]})


@router.post(
    "/conversations/{conv_id}/context",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def set_context(
    request: Request, conv_id: str, user: dict = Depends(require_user)
):
    """Set or update conversation context prompt."""
    _ensure_db()
    uid = _uid(user)
    body = await request.json()
    context_prompt = body.get("context_prompt", "")

    # Validate length
    if isinstance(context_prompt, str) and len(context_prompt) > 2000:
        return api_error(
            422, "Context prompt too long (max 2000 characters)", "validation_error"
        )

    # Normalize empty → NULL
    context_prompt = context_prompt.strip() if context_prompt else None
    context_prompt = context_prompt or None

    with _db.acquire_sync() as conn:
        row = conn.execute(
            "SELECT 1 FROM chat.webchat_conversations "
            "WHERE id = %s AND unified_id = %s",
            (conv_id, uid),
        ).fetchone()
        if not row:
            return api_error(404, "Not found", "not_found")

        conn.execute(
            "UPDATE chat.webchat_conversations SET context_prompt = %s WHERE id = %s",
            (context_prompt, conv_id),
        )
        conn.commit()
    return api_ok(data={"context_prompt": context_prompt})


# --- Participants ---


@router.get(
    "/conversations/{conv_id}/participants",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def list_participants(
    request: Request, conv_id: str, user: dict = Depends(require_user)
):
    """List participants of a conversation."""
    _ensure_db()
    uid = _uid(user)
    if not validate_conversation_access(conv_id, uid):
        return api_error(404, "Not found", "not_found")
    with _db.acquire_sync() as conn:
        rows = conn.execute(
            """SELECT p.unified_id, p.role, p.joined_at, u.display_name
               FROM chat.webchat_participants p
               LEFT JOIN app.users u ON u.username = p.unified_id
               WHERE p.conversation_id = %s
               ORDER BY p.joined_at""",
            (conv_id,),
        ).fetchall()
    from ui.routes.ws_chat import _active_connections

    participants = [
        {
            "uid": r["unified_id"],
            "role": r["role"],
            "display_name": r["display_name"] or r["unified_id"],
            "joined_at": r["joined_at"].isoformat() if r["joined_at"] else None,
            "online": r["unified_id"] in _active_connections,
        }
        for r in rows
    ]
    return api_ok(data={"participants": participants})


@router.post(
    "/conversations/{conv_id}/invite",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def invite_user(
    request: Request, conv_id: str, user: dict = Depends(require_user)
):
    """Invite a registered user to a conversation. Owner only."""
    _ensure_db()
    uid = _uid(user)
    if not is_conversation_owner(conv_id, uid):
        return api_error(403, "Only the owner can invite", "forbidden")

    body = await request.json()
    target = body.get("username", "").strip().lower()
    if not target:
        return api_error(400, "username required", "validation_error")
    if target == uid:
        return api_error(400, "Cannot invite yourself", "validation_error")

    # Check target user exists
    from ui.auth.database import AuthDatabase

    auth_db = AuthDatabase()
    target_user = auth_db.get_user_by_username(target)
    if not target_user:
        return api_error(404, f"User '{target}' not found", "not_found")

    # Check not already participant
    if validate_conversation_access(conv_id, target):
        return api_error(400, "User already in conversation", "validation_error")

    with _db.acquire_sync() as conn:
        # Convert private → shared on first invite
        conv = conn.execute(
            "SELECT type FROM chat.webchat_conversations WHERE id = %s",
            (conv_id,),
        ).fetchone()
        if conv and conv["type"] == "private":
            conn.execute(
                "UPDATE chat.webchat_conversations SET type = 'shared' WHERE id = %s",
                (conv_id,),
            )

        # Add participant
        conn.execute(
            """INSERT INTO chat.webchat_participants
               (conversation_id, unified_id, role)
               VALUES (%s, %s, 'member')
               ON CONFLICT DO NOTHING""",
            (conv_id, target),
        )
        conn.commit()

    # Notify via WebSocket
    from ui.routes.ws_chat import push_to_webchat

    await push_to_webchat(
        target,
        {
            "type": "conversation_invited",
            "conversation_id": conv_id,
        },
    )
    return api_ok(data={"invited": target})


@router.post(
    "/conversations/{conv_id}/invite-link",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def create_invite_link(
    request: Request, conv_id: str, user: dict = Depends(require_user)
):
    """Generate a shareable invite link. Owner only."""
    import secrets as stdlib_secrets

    _ensure_db()
    uid = _uid(user)
    if not is_conversation_owner(conv_id, uid):
        return api_error(403, "Only the owner can create invite links", "forbidden")

    body = await request.json()
    expires_hours = body.get("expires_hours", 72)
    max_uses = body.get("max_uses", 0)

    token = stdlib_secrets.token_urlsafe(32)
    with _db.acquire_sync() as conn:
        conn.execute(
            """INSERT INTO chat.webchat_invites
               (token, conversation_id, created_by, expires_at, max_uses)
               VALUES (%s, %s, %s, NOW() + INTERVAL '%s hours', %s)""",
            (token, conv_id, uid, expires_hours, max_uses),
        )
        # Convert to shared if private
        conn.execute(
            "UPDATE chat.webchat_conversations "
            "SET type = 'shared' WHERE id = %s AND type = 'private'",
            (conv_id,),
        )
        conn.commit()

    return api_ok(data={"url": f"/me/chat/join/{token}", "token": token})


@router.post(
    "/conversations/{conv_id}/leave",
    response_model=ApiResponse,
    response_model_exclude_none=True,
)
async def leave_conversation(
    request: Request, conv_id: str, user: dict = Depends(require_user)
):
    """Leave a shared conversation. Owner must transfer first."""
    _ensure_db()
    uid = _uid(user)
    role = validate_conversation_access(conv_id, uid)
    if not role:
        return api_error(404, "Not found", "not_found")
    if role == "owner":
        return api_error(
            400,
            "Owner must transfer ownership before leaving",
            "validation_error",
        )

    with _db.acquire_sync() as conn:
        conn.execute(
            "DELETE FROM chat.webchat_participants "
            "WHERE conversation_id = %s AND unified_id = %s",
            (conv_id, uid),
        )
        # Revert to private if only owner remains
        remaining = conn.execute(
            "SELECT count(*) as cnt FROM chat.webchat_participants "
            "WHERE conversation_id = %s",
            (conv_id,),
        ).fetchone()
        if remaining and remaining["cnt"] <= 1:
            conn.execute(
                "UPDATE chat.webchat_conversations SET type = 'private' WHERE id = %s",
                (conv_id,),
            )
        conn.commit()
    return api_ok()


@router.post(
    "/conversations/{conv_id}/transfer-ownership",
    response_model=ApiResponse,
    response_model_exclude_none=True,
)
async def transfer_ownership(
    request: Request, conv_id: str, user: dict = Depends(require_user)
):
    """Transfer conversation ownership to another participant. Owner only."""
    _ensure_db()
    uid = _uid(user)
    if not is_conversation_owner(conv_id, uid):
        return api_error(403, "Only the owner can transfer", "forbidden")

    body = await request.json()
    new_owner = body.get("new_owner", "").strip().lower()
    if not new_owner or new_owner == uid:
        return api_error(400, "Invalid new_owner", "validation_error")

    if not validate_conversation_access(conv_id, new_owner):
        return api_error(400, "Target is not a participant", "validation_error")

    with _db.acquire_sync() as conn:
        conn.execute(
            "UPDATE chat.webchat_participants SET role = 'member' "
            "WHERE conversation_id = %s AND unified_id = %s",
            (conv_id, uid),
        )
        conn.execute(
            "UPDATE chat.webchat_participants SET role = 'owner' "
            "WHERE conversation_id = %s AND unified_id = %s",
            (conv_id, new_owner),
        )
        conn.commit()
    return api_ok()


@router.post(
    "/conversations/{conv_id}/remove-participant",
    response_model=ApiResponse,
    response_model_exclude_none=True,
)
async def remove_participant(
    request: Request, conv_id: str, user: dict = Depends(require_user)
):
    """Remove a participant from the conversation. Owner only."""
    _ensure_db()
    uid = _uid(user)
    if not is_conversation_owner(conv_id, uid):
        return api_error(403, "Only the owner can remove participants", "forbidden")

    body = await request.json()
    target = body.get("username", "").strip().lower()
    if not target or target == uid:
        return api_error(400, "Invalid target", "validation_error")

    with _db.acquire_sync() as conn:
        conn.execute(
            "DELETE FROM chat.webchat_participants "
            "WHERE conversation_id = %s AND unified_id = %s",
            (conv_id, target),
        )
        # Revert to private if only owner remains
        remaining = conn.execute(
            "SELECT count(*) as cnt FROM chat.webchat_participants "
            "WHERE conversation_id = %s",
            (conv_id,),
        ).fetchone()
        if remaining and remaining["cnt"] <= 1:
            conn.execute(
                "UPDATE chat.webchat_conversations SET type = 'private' WHERE id = %s",
                (conv_id,),
            )
        conn.commit()

    from ui.routes.ws_chat import push_to_webchat

    await push_to_webchat(
        target, {"type": "participant_removed", "conversation_id": conv_id}
    )
    return api_ok()


# --- Pin / Unpin ---


@router.post(
    "/conversations/{conv_id}/pin",
    response_model=ApiResponse,
    response_model_exclude_none=True,
)
async def pin_conversation(conv_id: str, user: dict = Depends(require_user)):
    """Pin a conversation to the top of the list for this user."""
    _ensure_db()
    uid = _uid(user)
    if not validate_conversation_access(conv_id, uid):
        return api_error(404, "Not found", "not_found")

    with _db.acquire_sync() as conn:
        conn.execute(
            "UPDATE chat.webchat_participants SET pinned_at = NOW() "
            "WHERE conversation_id = %s AND unified_id = %s AND pinned_at IS NULL",
            (conv_id, uid),
        )
        conn.commit()
    return api_ok()


@router.post(
    "/conversations/{conv_id}/unpin",
    response_model=ApiResponse,
    response_model_exclude_none=True,
)
async def unpin_conversation(conv_id: str, user: dict = Depends(require_user)):
    """Unpin a conversation."""
    _ensure_db()
    uid = _uid(user)
    if not validate_conversation_access(conv_id, uid):
        return api_error(404, "Not found", "not_found")

    with _db.acquire_sync() as conn:
        conn.execute(
            "UPDATE chat.webchat_participants SET pinned_at = NULL "
            "WHERE conversation_id = %s AND unified_id = %s",
            (conv_id, uid),
        )
        conn.commit()
    return api_ok()


# --- File upload ---


@router.post(
    "/upload",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def upload_file(request: Request, user: dict = Depends(require_user)):
    """Upload a file for chat attachment. Returns the server-side file path."""
    uid = _uid(user)

    form = await request.form()
    file: UploadFile | None = form.get("file")
    if not file or not file.filename:
        return api_error(400, "No file provided", "validation_error")

    # Validate extension
    from handlers.attachment_handler import sanitize_filename

    safe_name = sanitize_filename(file.filename)
    ext = Path(safe_name).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return api_error(400, f"File type {ext} not allowed", "validation_error")

    # Read with size limit
    content = await file.read()
    if len(content) > MAX_UPLOAD_SIZE:
        return api_error(400, "File too large (max 20MB)", "validation_error")

    # Save to user-specific directory with timestamp prefix to avoid collisions
    user_dir = ATTACHMENTS_DIR / "webchat" / uid
    user_dir.mkdir(parents=True, exist_ok=True)
    dest_name = f"{int(time.time())}_{safe_name}"
    dest_path = user_dir / dest_name
    dest_path.write_bytes(content)

    logger.info(f"WebChat upload: {uid} -> {dest_path} ({len(content)} bytes)")

    return api_ok(
        data={
            "path": str(dest_path),
            "filename": safe_name,
            "size": len(content),
        }
    )


# --- Bot file delivery ---

OUTBOUND_DIR = ATTACHMENTS_DIR / "webchat-outbound"

# Token store: {token: {"path": str, "uid": str, "expires": float}}
_file_tokens: dict[str, dict] = {}
_FILE_TOKEN_TTL = 3600  # 1 hour


def create_file_token(
    file_path: str, uid: str, conversation_id: str | None = None
) -> str:
    """Create a short-lived token to serve a file to a specific user."""
    import secrets

    token = secrets.token_urlsafe(32)
    _file_tokens[token] = {
        "path": file_path,
        "uid": uid,
        "conversation_id": conversation_id,
        "expires": time.time() + _FILE_TOKEN_TTL,
    }
    # Prune expired tokens periodically (every 100 creates)
    if len(_file_tokens) % 100 == 0:
        now = time.time()
        expired = [k for k, v in _file_tokens.items() if v["expires"] < now]
        for k in expired:
            del _file_tokens[k]
    return token


@router.get("/files/{token}")
async def serve_file(token: str, user: dict = Depends(require_user)):
    """Serve a bot-delivered file via token-based access."""
    entry = _file_tokens.get(token)
    if not entry:
        return api_error(404, "File not found or expired", "not_found")

    if entry["expires"] < time.time():
        _file_tokens.pop(token, None)
        return api_error(404, "File expired", "not_found")

    uid = _uid(user)
    if entry["uid"] != uid:
        # Allow access if user is a participant of the same conversation
        conv_id = entry.get("conversation_id")
        if not conv_id or not validate_conversation_access(conv_id, uid):
            return api_error(403, "Access denied", "forbidden")

    file_path = Path(entry["path"])
    if not file_path.exists():
        _file_tokens.pop(token, None)
        return api_error(404, "File no longer available", "not_found")

    import mimetypes

    content_type = mimetypes.guess_type(str(file_path))[0] or "application/octet-stream"
    return FileResponse(file_path, media_type=content_type)


# --- Conversation documents ---

DOCS_DIR = ATTACHMENTS_DIR / "webchat-docs"


@router.get(
    "/conversations/{conv_id}/documents",
    response_model=ApiResponse[list],
    response_model_exclude_none=True,
)
async def list_documents(conv_id: str, user: dict = Depends(require_user)):
    """List documents attached to a conversation."""
    _ensure_db()
    uid = _uid(user)
    if not validate_conversation_access(conv_id, uid):
        return api_error(403, "Access denied", "forbidden")

    with _db.acquire_sync() as conn:
        rows = conn.execute(
            "SELECT id, original_filename, file_size, mime_type, "
            "uploaded_by, uploaded_at "
            "FROM chat.webchat_documents "
            "WHERE conversation_id = %s ORDER BY uploaded_at",
            (conv_id,),
        ).fetchall()
    return api_ok(data=[dict(r) for r in rows])


@router.post(
    "/conversations/{conv_id}/documents",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def upload_document(
    conv_id: str, request: Request, user: dict = Depends(require_user)
):
    """Upload a document to a conversation's knowledge base."""
    _ensure_db()
    uid = _uid(user)
    if not validate_conversation_access(conv_id, uid):
        return api_error(403, "Access denied", "forbidden")

    form = await request.form()
    file: UploadFile | None = form.get("file")
    if not file or not file.filename:
        return api_error(400, "No file provided", "validation_error")

    from handlers.attachment_handler import sanitize_filename

    safe_name = sanitize_filename(file.filename)
    ext = Path(safe_name).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return api_error(400, f"File type {ext} not allowed", "validation_error")

    content = await file.read()
    if len(content) > MAX_UPLOAD_SIZE:
        return api_error(400, "File too large (max 20MB)", "validation_error")

    # Save to conversation-specific directory
    conv_dir = DOCS_DIR / conv_id
    conv_dir.mkdir(parents=True, exist_ok=True)
    doc_id = str(uuid.uuid4())
    dest_name = f"{doc_id}_{safe_name}"
    dest_path = conv_dir / dest_name
    dest_path.write_bytes(content)

    # Extract text
    mime = file.content_type or ""
    content_text = _extract_text(str(dest_path), mime)

    with _db.acquire_sync() as conn:
        conn.execute(
            "INSERT INTO chat.webchat_documents "
            "(id, conversation_id, filename, original_filename, file_path, "
            "file_size, mime_type, content_text, uploaded_by) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (
                doc_id,
                conv_id,
                dest_name,
                safe_name,
                str(dest_path),
                len(content),
                mime,
                content_text,
                uid,
            ),
        )
        conn.commit()

    logger.info(
        f"WebChat doc upload: {uid} -> {conv_id[:8]}... "
        f"{safe_name} ({len(content)} bytes, text={len(content_text)} chars)"
    )

    return api_ok(
        data={
            "id": doc_id,
            "original_filename": safe_name,
            "file_size": len(content),
            "mime_type": mime,
            "uploaded_by": uid,
            "has_text": bool(content_text),
        }
    )


@router.delete(
    "/conversations/{conv_id}/documents/{doc_id}",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def delete_document(
    conv_id: str, doc_id: str, user: dict = Depends(require_user)
):
    """Remove a document. Available to the uploader or conversation owner."""
    _ensure_db()
    uid = _uid(user)

    with _db.acquire_sync() as conn:
        row = conn.execute(
            "SELECT uploaded_by, file_path FROM chat.webchat_documents "
            "WHERE id = %s AND conversation_id = %s",
            (doc_id, conv_id),
        ).fetchone()
        if not row:
            return api_error(404, "Document not found", "not_found")

        # Only uploader or owner can delete
        if row["uploaded_by"] != uid and not is_conversation_owner(conv_id, uid):
            return api_error(403, "Not authorized to delete", "forbidden")

        conn.execute("DELETE FROM chat.webchat_documents WHERE id = %s", (doc_id,))
        conn.commit()

    # Remove file from disk
    try:
        file_path = Path(row["file_path"])
        if file_path.exists():
            file_path.unlink()
    except Exception:
        pass

    return api_ok()


@router.get("/conversations/{conv_id}/documents/{doc_id}/download")
async def download_document(
    conv_id: str, doc_id: str, user: dict = Depends(require_user)
):
    """Download a conversation document. Any participant can download."""
    _ensure_db()
    uid = _uid(user)
    if not validate_conversation_access(conv_id, uid):
        return api_error(403, "Access denied", "forbidden")

    with _db.acquire_sync() as conn:
        row = conn.execute(
            "SELECT original_filename, file_path, mime_type "
            "FROM chat.webchat_documents "
            "WHERE id = %s AND conversation_id = %s",
            (doc_id, conv_id),
        ).fetchone()

    if not row:
        return api_error(404, "Document not found", "not_found")

    file_path = Path(row["file_path"])
    if not file_path.exists():
        return api_error(404, "File no longer available", "not_found")

    return FileResponse(
        file_path,
        media_type=row["mime_type"] or "application/octet-stream",
        filename=row["original_filename"],
    )


def get_conversation_documents(conversation_id: str) -> list[dict]:
    """Fetch documents with extracted text for context injection."""
    _ensure_db()
    with _db.acquire_sync() as conn:
        rows = conn.execute(
            "SELECT original_filename, content_text "
            "FROM chat.webchat_documents "
            "WHERE conversation_id = %s ORDER BY uploaded_at",
            (conversation_id,),
        ).fetchall()
    return [dict(r) for r in rows]


# --- Plan management ---


def get_active_plan(conversation_id: str) -> dict | None:
    """Get the active/draft/paused plan for a conversation, with tasks."""
    _ensure_db()
    with _db.acquire_sync() as conn:
        plan = conn.execute(
            "SELECT id, title, status, created_by, auto_continue_count, "
            "created_at, updated_at "
            "FROM chat.webchat_plans "
            "WHERE conversation_id = %s "
            "AND status NOT IN ('completed', 'cancelled') "
            "ORDER BY created_at DESC LIMIT 1",
            (conversation_id,),
        ).fetchone()
        if not plan:
            return None
        tasks = conn.execute(
            "SELECT id, position, title, description, status, result, "
            "started_at, completed_at "
            "FROM chat.webchat_plan_tasks "
            "WHERE plan_id = %s ORDER BY position",
            (plan["id"],),
        ).fetchall()
    return {**dict(plan), "tasks": [dict(t) for t in tasks]}


@router.get(
    "/conversations/{conv_id}/plan",
    response_model=ApiResponse[dict | None],
    response_model_exclude_none=True,
)
async def get_plan(conv_id: str, user: dict = Depends(require_user)):
    """Get current plan for a conversation."""
    _ensure_db()
    uid = _uid(user)
    if not validate_conversation_access(conv_id, uid):
        return api_error(403, "Access denied", "forbidden")
    plan = get_active_plan(conv_id)
    return api_ok(data=plan)


@router.post(
    "/conversations/{conv_id}/plan/tasks",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def add_plan_task(
    conv_id: str, request: Request, user: dict = Depends(require_user)
):
    """Add a task to the current plan. Only when draft/paused."""
    _ensure_db()
    uid = _uid(user)
    if not validate_conversation_access(conv_id, uid):
        return api_error(403, "Access denied", "forbidden")

    body = await request.json()
    title = body.get("title", "").strip()
    if not title:
        return api_error(400, "Title required", "validation_error")

    with _db.acquire_sync() as conn:
        plan = conn.execute(
            "SELECT id, status FROM chat.webchat_plans "
            "WHERE conversation_id = %s "
            "AND status NOT IN ('completed', 'cancelled') LIMIT 1",
            (conv_id,),
        ).fetchone()
        if not plan:
            return api_error(404, "No active plan", "not_found")
        if plan["status"] not in ("draft", "paused"):
            return api_error(
                400, "Plan must be draft or paused to add tasks", "invalid_state"
            )

        max_pos = conn.execute(
            "SELECT COALESCE(MAX(position), -1) as mp "
            "FROM chat.webchat_plan_tasks WHERE plan_id = %s",
            (plan["id"],),
        ).fetchone()["mp"]

        task_id = str(uuid.uuid4())
        conn.execute(
            "INSERT INTO chat.webchat_plan_tasks "
            "(id, plan_id, position, title, description, status) "
            "VALUES (%s, %s, %s, %s, %s, 'pending')",
            (task_id, plan["id"], max_pos + 1, title, body.get("description")),
        )
        conn.commit()

    from ui.routes.ws_chat import broadcast_to_conversation

    await broadcast_to_conversation(
        conv_id,
        {
            "type": "plan_updated",
            "plan_id": plan["id"],
            "changes": {
                "added_tasks": [
                    {
                        "id": task_id,
                        "position": max_pos + 1,
                        "title": title,
                        "status": "pending",
                    }
                ]
            },
        },
    )
    return api_ok(data={"id": task_id, "title": title})


@router.delete(
    "/conversations/{conv_id}/plan/tasks/{task_id}",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def remove_plan_task(
    conv_id: str, task_id: str, user: dict = Depends(require_user)
):
    """Remove a pending task. Only when draft/paused."""
    _ensure_db()
    uid = _uid(user)
    if not validate_conversation_access(conv_id, uid):
        return api_error(403, "Access denied", "forbidden")

    with _db.acquire_sync() as conn:
        plan = conn.execute(
            "SELECT id, status FROM chat.webchat_plans "
            "WHERE conversation_id = %s "
            "AND status NOT IN ('completed', 'cancelled') LIMIT 1",
            (conv_id,),
        ).fetchone()
        if not plan or plan["status"] not in ("draft", "paused"):
            return api_error(400, "Plan must be draft or paused", "invalid_state")

        deleted = conn.execute(
            "DELETE FROM chat.webchat_plan_tasks "
            "WHERE id = %s AND plan_id = %s AND status = 'pending' RETURNING id",
            (task_id, plan["id"]),
        ).fetchone()
        if not deleted:
            return api_error(404, "Task not found or not pending", "not_found")
        conn.commit()

    from ui.routes.ws_chat import broadcast_to_conversation

    await broadcast_to_conversation(
        conv_id,
        {
            "type": "plan_updated",
            "plan_id": plan["id"],
            "changes": {"removed_tasks": [task_id]},
        },
    )
    return api_ok()


@router.post(
    "/conversations/{conv_id}/plan/tasks/{task_id}",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def update_plan_task(
    conv_id: str,
    task_id: str,
    request: Request,
    user: dict = Depends(require_user),
):
    """Update a task's title/description. Only when draft/paused."""
    _ensure_db()
    uid = _uid(user)
    if not validate_conversation_access(conv_id, uid):
        return api_error(403, "Access denied", "forbidden")

    body = await request.json()
    title = body.get("title", "").strip()
    description = body.get("description")

    if not title:
        return api_error(400, "Title required", "validation_error")

    with _db.acquire_sync() as conn:
        plan = conn.execute(
            "SELECT id, status FROM chat.webchat_plans "
            "WHERE conversation_id = %s "
            "AND status NOT IN ('completed', 'cancelled') LIMIT 1",
            (conv_id,),
        ).fetchone()
        if not plan or plan["status"] not in ("draft", "paused"):
            return api_error(400, "Plan must be draft or paused", "invalid_state")

        conn.execute(
            "UPDATE chat.webchat_plan_tasks SET title = %s, description = %s "
            "WHERE id = %s AND plan_id = %s",
            (title, description, task_id, plan["id"]),
        )
        conn.commit()

    return api_ok()


@router.post(
    "/conversations/{conv_id}/plan/tasks/reorder",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def reorder_plan_tasks(
    conv_id: str, request: Request, user: dict = Depends(require_user)
):
    """Reorder tasks. Only when draft/paused."""
    _ensure_db()
    uid = _uid(user)
    if not validate_conversation_access(conv_id, uid):
        return api_error(403, "Access denied", "forbidden")

    body = await request.json()
    task_ids = body.get("task_ids", [])
    if not task_ids:
        return api_error(400, "task_ids required", "validation_error")

    with _db.acquire_sync() as conn:
        plan = conn.execute(
            "SELECT id, status FROM chat.webchat_plans "
            "WHERE conversation_id = %s "
            "AND status NOT IN ('completed', 'cancelled') LIMIT 1",
            (conv_id,),
        ).fetchone()
        if not plan or plan["status"] not in ("draft", "paused"):
            return api_error(400, "Plan must be draft or paused", "invalid_state")

        for i, tid in enumerate(task_ids):
            conn.execute(
                "UPDATE chat.webchat_plan_tasks SET position = %s "
                "WHERE id = %s AND plan_id = %s",
                (i, tid, plan["id"]),
            )
        conn.commit()

    from ui.routes.ws_chat import broadcast_to_conversation

    await broadcast_to_conversation(
        conv_id,
        {
            "type": "plan_updated",
            "plan_id": plan["id"],
            "changes": {"reordered": task_ids},
        },
    )
    return api_ok()


@router.post(
    "/conversations/{conv_id}/plan/status",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def update_plan_status(
    conv_id: str, request: Request, user: dict = Depends(require_user)
):
    """User-initiated plan status change (pause/cancel from panel)."""
    _ensure_db()
    uid = _uid(user)
    if not validate_conversation_access(conv_id, uid):
        return api_error(403, "Access denied", "forbidden")

    body = await request.json()
    new_status = body.get("status", "").strip()
    if new_status not in ("paused", "cancelled"):
        return api_error(400, "Status must be paused or cancelled", "validation_error")

    with _db.acquire_sync() as conn:
        plan = conn.execute(
            "SELECT id, status FROM chat.webchat_plans "
            "WHERE conversation_id = %s "
            "AND status NOT IN ('completed', 'cancelled') LIMIT 1",
            (conv_id,),
        ).fetchone()
        if not plan:
            return api_error(404, "No active plan", "not_found")

        if new_status == "paused" and plan["status"] != "active":
            return api_error(400, "Can only pause an active plan", "invalid_state")

        conn.execute(
            "UPDATE chat.webchat_plans SET status = %s, "
            "updated_at = CURRENT_TIMESTAMP WHERE id = %s",
            (new_status, plan["id"]),
        )
        conn.commit()

    from ui.routes.ws_chat import broadcast_to_conversation

    await broadcast_to_conversation(
        conv_id,
        {
            "type": "plan_updated",
            "plan_id": plan["id"],
            "changes": {"status": new_status},
        },
    )
    return api_ok()


@router.post(
    "/conversations/{conv_id}/plan/resume",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def resume_plan(conv_id: str, user: dict = Depends(require_user)):
    """Resume a paused plan. Resets auto_continue_count and sets active."""
    _ensure_db()
    uid = _uid(user)
    if not validate_conversation_access(conv_id, uid):
        return api_error(403, "Access denied", "forbidden")

    with _db.acquire_sync() as conn:
        plan = conn.execute(
            "SELECT id, status FROM chat.webchat_plans "
            "WHERE conversation_id = %s AND status = 'paused' LIMIT 1",
            (conv_id,),
        ).fetchone()
        if not plan:
            return api_error(404, "No paused plan", "not_found")

        conn.execute(
            "UPDATE chat.webchat_plans SET status = 'active', "
            "auto_continue_count = 0, updated_at = CURRENT_TIMESTAMP "
            "WHERE id = %s",
            (plan["id"],),
        )
        conn.commit()

    from ui.routes.ws_chat import broadcast_to_conversation

    await broadcast_to_conversation(
        conv_id,
        {
            "type": "plan_updated",
            "plan_id": plan["id"],
            "changes": {"status": "active"},
        },
    )
    return api_ok()


# --- TTS ---


def _strip_markdown(text: str) -> str:
    """Remove markdown formatting for cleaner TTS input."""
    text = re.sub(r"```[\s\S]*?```", " code block ", text)
    text = re.sub(r"`[^`]+`", lambda m: m.group(0)[1:-1], text)
    text = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)
    text = re.sub(r"[#*_~>|]", "", text)
    text = re.sub(r"\n{2,}", ". ", text)
    text = text.replace("\n", " ").strip()
    return text


@router.get(
    "/tts/config",
    response_model=ApiResponse[dict],
    response_model_exclude_none=True,
)
async def tts_config(user: dict = Depends(require_user)):
    """Return the current TTS provider setting."""
    from ui.config_manager import ConfigManager

    config = ConfigManager()
    return api_ok(data={"provider": config.get_webchat_tts_provider()})


@router.post("/tts")
async def tts_synthesize(request: Request, user: dict = Depends(require_user)):
    """Synthesize text to speech and return audio file."""
    from ui import tts_service
    from ui.config_manager import ConfigManager

    try:
        body = await request.json()
    except Exception:
        return api_error(400, "Invalid JSON", "validation_error")

    text = (body.get("text") or "").strip()
    if not text:
        return api_error(400, "No text provided", "validation_error")

    clean_text = _strip_markdown(text)
    if not clean_text:
        return api_error(400, "No speakable text", "validation_error")

    config = ConfigManager()
    provider = config.get_webchat_tts_provider()

    try:
        file_path = await tts_service.synthesize(clean_text, provider, locale="it")
    except Exception as e:
        logger.error(f"TTS synthesis error ({provider}): {e}")
        return api_error(500, "TTS synthesis failed", "internal_error")

    if not file_path:
        return api_error(400, "Provider is browser", "browser_tts")

    return FileResponse(
        file_path,
        media_type="audio/mpeg",
        background=BackgroundTask(os.unlink, file_path),
    )
