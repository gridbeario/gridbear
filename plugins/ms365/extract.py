"""Pure helpers for the ms365 MCP server: Office/PDF text extraction and
Microsoft Graph sharing-URL encoding. No plugin state, no I/O."""

import base64
import io
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def encode_sharing_url(url: str) -> str:
    """Encode a sharing URL into a Graph shareId (already `u!`-prefixed).

    Callers use it as `/shares/{enc}/driveItem` — NOT `/shares/u!{enc}` (that
    would double the prefix to `u!u!…` and 400).
    """
    b64 = base64.b64encode(url.encode("utf-8")).decode("ascii")
    return "u!" + b64.rstrip("=").replace("/", "_").replace("+", "-")


_TEXT_EXTS = (".txt", ".md", ".csv", ".json", ".xml")


def _truncate(text: str, max_chars: int) -> str:
    if len(text) > max_chars:
        return text[:max_chars] + "\n...(truncated)"
    return text


def extract_text(content: bytes, filename: str, max_chars: int = 200_000) -> str:
    """Extract readable text from document bytes, sniffing by extension.

    Never raises: parse errors and unsupported/empty inputs return a readable
    note. `filename` is only used for the extension; `content` is the bytes.
    """
    ext = Path(filename).suffix.lower()

    if ext in _TEXT_EXTS:
        return _truncate(content.decode("utf-8", errors="replace"), max_chars)

    if ext not in (".pdf", ".docx", ".xlsx", ".xls"):
        return f"unsupported format ({ext or 'no extension'}) — cannot extract text"

    bio = io.BytesIO(content)
    try:
        if ext == ".pdf":
            from pypdf import PdfReader

            reader = PdfReader(bio)
            pages = [p.extract_text() for p in reader.pages]
            text = "\n\n".join(p for p in pages if p)
        elif ext == ".docx":
            from docx import Document

            doc = Document(bio)
            text = "\n\n".join(p.text for p in doc.paragraphs if p.text)
        else:  # .xlsx / .xls
            from openpyxl import load_workbook

            wb = load_workbook(bio, read_only=True, data_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"[Sheet: {sheet}]")
                for row in ws.iter_rows(values_only=True):
                    lines.append("\t".join("" if c is None else str(c) for c in row))
            wb.close()
            text = "\n".join(lines)
    except Exception as err:
        logger.warning("ms365 extract_text failed for %s: %s", filename, err)
        return f"could not parse {filename}: {type(err).__name__}: {err}"

    if not text or not text.strip():
        return "no extractable text (document may be image-only/scanned or empty)"
    return _truncate(text, max_chars)
